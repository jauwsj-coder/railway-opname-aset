import os
import io
import unittest
from unittest.mock import patch

import app as application
from openpyxl import load_workbook


def values_for(headers, data):
    return [data.get(header, "") for header in headers]


class FakeWorksheet:
    def __init__(self, title, values):
        self.title, self.values, self.appended = title, values, []

    def get_all_values(self): return self.values
    def batch_update(self, updates, value_input_option=None): self.updates = updates
    def append_row(self, row, value_input_option=None): self.appended.append(row); self.values.append(row)
    def clear(self): self.values = []
    def update(self, values, range_name=None, value_input_option=None):
        if range_name == "A1" or not self.values:
            self.values = values
        else:
            self.values[0].extend(values[0])
    def format(self, *args, **kwargs): pass
    def freeze(self, *args, **kwargs): pass


class FakeSpreadsheet:
    def __init__(self, worksheets): self.worksheets = worksheets
    def worksheet(self, name):
        if name not in self.worksheets: raise application.gspread.WorksheetNotFound(name)
        return self.worksheets[name]
    def add_worksheet(self, title, rows, cols):
        self.worksheets[title] = FakeWorksheet(title, [])
        return self.worksheets[title]


class AppTest(unittest.TestCase):
    def setUp(self):
        os.environ["APP_SECRET_KEY"] = "test-secret"
        os.environ["PHOTO_UPLOAD_SCRIPT_URL"] = "https://script.google.com/test"
        os.environ["PHOTO_UPLOAD_SECRET"] = "photo-secret"
        self.master = FakeWorksheet("MASTER_ASET", [application.MASTER_HEADERS, ["AST-0001", "LAPTOP", "LT-01", "BUDI", "DONE", "OK", "KANTOR", "AREA A", "", "", "", ""]])
        self.log = FakeWorksheet("LOG_OPNAME", [application.LOG_HEADERS])
        self.role = FakeWorksheet("ROLE", [
            application.ROLE_HEADERS,
            ["ADMIN", "1001", " SUPER ADMIN ", " all ", ""],
            ["ADMIN PIC", "1002", "SUPER ADMIN, PIC ASET", "ALL", "AREA A"],
            ["PIC MULTI", "1003", "PIC ASET", "AREA B, AREA A", ""],
            ["INVALID", "1004", "ADMIN", "ALL", ""],
            ["PIC EMPTY", "1005", "PIC ASET", "AREA Z", ""],
            ["GA CORPORATE", "1006", "SUPER ADMIN", "ALL", ""],
        ])
        self.dashboard = FakeWorksheet("DASHBOARD", [application.DASHBOARD_HEADERS])
        self.approval = FakeWorksheet("APPROVAL_PERUBAHAN_ASET", [application.APPROVAL_HEADERS])
        self.sheets = {"MASTER_ASET": self.master, "LOG_OPNAME": self.log, "ROLE": self.role, "DASHBOARD": self.dashboard, "APPROVAL_PERUBAHAN_ASET": self.approval}
        self.client = application.app.test_client()

    def worksheet(self, name): return self.sheets[name]

    def login(self, name="ADMIN PIC", user_id="1002"):
        response = self.client.post("/api/login", json={"name": name, "userId": user_id})
        return response, {"Authorization": "Bearer " + response.json["token"]} if response.status_code == 200 else {}

    @patch("app.get_worksheet")
    def test_login_normalization_validation_and_area_error(self, get_worksheet):
        get_worksheet.side_effect = self.worksheet
        response, _ = self.login(" admin ", " 1001 ")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json["user"]["role"], "SUPER ADMIN")
        self.assertEqual(response.json["user"]["area"], "ALL")
        self.assertEqual(self.login("INVALID", "1004")[0].status_code, 403)
        self.assertEqual(self.login("ADMIN", "WRONG")[0].status_code, 401)

    @patch("app.get_worksheet")
    def test_submit_log_dashboard_and_score_rules(self, get_worksheet):
        get_worksheet.side_effect = self.worksheet
        _, headers = self.login()
        response = self.client.post("/api/opname", headers=headers, json={"assetCode": "AST-0001", "condition": "BAIK", "notes": "Sesuai"})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(self.log.appended[0]), len(application.LOG_HEADERS))
        self.assertEqual(self.log.appended[0][-3:], ["ADMIN PIC", "1002", "SUPER ADMIN, PIC ASET"])
        self.assertEqual(response.json["summary"], {"total": 1, "completed": 1, "pending": 0, "good": 1, "damaged": 0})
        admin_score = next(item for item in response.json["scoreCard"] if item["name"] == "ADMIN PIC")
        self.assertEqual(admin_score["progress"], 100)
        self.assertEqual(admin_score["completed"], 1)
        self.assertEqual(admin_score["status"], "Selesai")

        second = self.client.post("/api/opname", headers=headers, json={"assetCode": "AST-0001", "condition": "RUSAK", "notes": "Perlu perbaikan"})
        self.assertEqual(second.json["summary"], {"total": 1, "completed": 1, "pending": 0, "good": 0, "damaged": 1})
        admin_score = next(item for item in second.json["scoreCard"] if item["name"] == "ADMIN PIC")
        self.assertEqual(admin_score["completed"], 1)
        self.assertEqual(admin_score["progress"], 100)

        self.log.values = [application.LOG_HEADERS]
        dashboard = self.client.get("/api/dashboard", headers=headers).json
        self.assertEqual(dashboard["summary"], {"total": 1, "completed": 0, "pending": 1, "good": 0, "damaged": 0})
        self.assertTrue(all(item["progress"] == 0 for item in dashboard["scoreCard"]))

    @patch("app.get_worksheet")
    def test_pure_super_admin_excluded_from_score(self, get_worksheet):
        get_worksheet.side_effect = self.worksheet
        _, headers = self.login("ADMIN", "1001")
        self.client.post("/api/opname", headers=headers, json={"assetCode": "AST-0001", "condition": "RUSAK", "notes": ""})
        dashboard = self.client.get("/api/dashboard", headers=headers).json
        self.assertEqual(dashboard["summary"]["damaged"], 1)
        self.assertFalse(any(item["name"] == "ADMIN" for item in dashboard["scoreCard"]))

    @patch("app.get_worksheet")
    def test_pic_without_matching_area_rejected(self, get_worksheet):
        get_worksheet.side_effect = self.worksheet
        response, _ = self.login("PIC EMPTY", "1005")
        self.assertEqual(response.status_code, 403)
        self.assertIn("tidak memiliki aset", response.json["message"])

    @patch("app.get_worksheet")
    def test_multi_area_access_and_scorecard_fallback(self, get_worksheet):
        get_worksheet.side_effect = self.worksheet
        response, headers = self.login("PIC MULTI", "1003")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json["user"]["areas"], ["AREA A", "AREA B"])
        dashboard = self.client.get("/api/dashboard", headers=headers).json
        score = next(item for item in dashboard["scoreCard"] if item["name"] == "PIC MULTI")
        self.assertEqual(score["scoreAreas"], "AREA B, AREA A")
        self.assertEqual(score["total"], 1)
        self.assertEqual(score["progress"], 0)

    @patch("app.get_worksheet")
    def test_pic_change_request_waits_for_ga_corporate_approval(self, get_worksheet):
        get_worksheet.side_effect = self.worksheet
        _, pic_headers = self.login("PIC MULTI", "1003")
        response = self.client.post("/api/asset-change-requests", headers=pic_headers, json={
            "assetCode": "AST-0001", "user": "USER BARU", "area": "AREA B",
            "detailLocation": "RUANG BARU", "reason": "Aset dipindahkan",
        })
        self.assertEqual(response.status_code, 200)
        request_row = dict(zip(application.APPROVAL_HEADERS, self.approval.appended[0]))
        self.assertEqual(request_row["STATUS APPROVAL"], "PENDING")
        self.assertEqual(request_row["USER LAMA"], "BUDI")
        self.assertFalse(hasattr(self.master, "updates"))

        denied = self.client.post(f"/api/asset-change-requests/{response.json['requestId']}/approve", headers=pic_headers, json={})
        self.assertEqual(denied.status_code, 403)

        _, ga_headers = self.login("GA CORPORATE", "1006")
        approved = self.client.post(f"/api/asset-change-requests/{response.json['requestId']}/approve", headers=ga_headers, json={"notes": "Sesuai"})
        self.assertEqual(approved.status_code, 200)
        master_updates = {item["range"]: item["values"][0][0] for item in self.master.updates}
        self.assertEqual(master_updates["D2"], "USER BARU")
        self.assertEqual(master_updates["H2"], "AREA B")
        self.assertEqual(master_updates["G2"], "RUANG BARU")
        approval_updates = {item["range"]: item["values"][0][0] for item in self.approval.updates}
        self.assertIn("APPROVED", approval_updates.values())
        self.assertIn("GA CORPORATE", approval_updates.values())

    @patch("app.get_worksheet")
    def test_asset_change_request_rejects_unchanged_and_duplicate_pending(self, get_worksheet):
        get_worksheet.side_effect = self.worksheet
        _, headers = self.login("PIC MULTI", "1003")
        unchanged = self.client.post("/api/asset-change-requests", headers=headers, json={
            "assetCode": "AST-0001", "user": "BUDI", "area": "AREA A", "detailLocation": "KANTOR", "reason": "Tes",
        })
        self.assertEqual(unchanged.status_code, 400)
        payload = {"assetCode": "AST-0001", "user": "BUDI BARU", "area": "AREA A", "detailLocation": "KANTOR", "reason": "Tes"}
        self.assertEqual(self.client.post("/api/asset-change-requests", headers=headers, json=payload).status_code, 200)
        duplicate = self.client.post("/api/asset-change-requests", headers=headers, json=payload)
        self.assertEqual(duplicate.status_code, 409)

    @patch("app.get_spreadsheet")
    def test_setup_appends_missing_headers_without_overwriting_data(self, get_spreadsheet):
        old_master = FakeWorksheet("MASTER_ASET", [["NOMOR ASSET", "TYPE"], ["AST-X", "MEJA"]])
        spreadsheet = FakeSpreadsheet({"MASTER_ASET": old_master})
        get_spreadsheet.return_value = spreadsheet
        os.environ["SETUP_TOKEN"] = "setup-secret"
        response = self.client.post("/api/setup", headers={"X-Setup-Token": "setup-secret"})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(old_master.values[1], ["AST-X", "MEJA"])
        self.assertTrue(all(header in old_master.values[0] for header in application.MASTER_HEADERS))

    def test_append_record_follows_actual_header_order(self):
        headers = ["TYPE", "NOMOR ASSET", *[h for h in application.LOG_HEADERS if h not in {"TYPE", "NOMOR ASSET"}]]
        sheet = FakeWorksheet("LOG_OPNAME", [headers])
        record = {header: header + "-VALUE" for header in application.LOG_HEADERS}
        application.append_record(sheet, "LOG_OPNAME", application.LOG_HEADERS, record)
        self.assertEqual(sheet.appended[0][0], "TYPE-VALUE")
        self.assertEqual(sheet.appended[0][1], "NOMOR ASSET-VALUE")

    @patch("app.call_photo_upload_script", return_value={"url": "https://drive.google.com/file/d/FILE-ID/view", "period": "2026-Jan-Jun"})
    @patch("app.get_worksheet")
    def test_upload_documentation_uses_apps_script_relay(self, get_worksheet, relay):
        get_worksheet.side_effect = self.worksheet
        _, headers = self.login()
        response = self.client.post(
            "/api/upload-documentation",
            headers=headers,
            data={"assetCode": "AST-0001", "photo": (io.BytesIO(b"photo"), "asset.jpg", "image/jpeg")},
            content_type="multipart/form-data",
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("drive.google.com", response.json["url"])
        relay.assert_called_once()
        self.assertEqual(relay.call_args.args[0]["action"], "upload")
        self.assertTrue(relay.call_args.args[0]["base64Data"])

    @patch("app.call_photo_upload_script", return_value={"cleanup": {"keptPeriods": ["2026-Jan-Jun"], "trashedFolders": []}})
    def test_cleanup_admin_endpoint_uses_apps_script(self, relay):
        os.environ["SETUP_TOKEN"] = "setup-secret"
        response = self.client.get("/api/cleanup-drive-photos?token=setup-secret")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json["cleanup"]["keptPeriods"], ["2026-Jan-Jun"])
        relay.assert_called_once_with({"action": "cleanup"})

    @patch("app.call_photo_upload_script", return_value={"message": "Tes berhasil", "testFileMovedToTrash": True})
    def test_photo_upload_admin_endpoint(self, relay):
        os.environ["SETUP_TOKEN"] = "setup-secret"
        response = self.client.get("/api/test-photo-upload?token=setup-secret")
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json["testFileMovedToTrash"])
        self.assertEqual(relay.call_args.args[0]["action"], "test")

    @patch("app.get_worksheet")
    def test_empty_photo_is_allowed(self, get_worksheet):
        get_worksheet.side_effect = self.worksheet
        _, headers = self.login()
        response = self.client.post("/api/upload-documentation", headers=headers, data={"assetCode": "AST-0001"})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json["url"], "")

    def test_unknown_endpoint_returns_clean_json_404(self):
        response = self.client.get("/not-a-real-endpoint")
        self.assertEqual(response.status_code, 404)
        self.assertEqual(response.json["path"], "/not-a-real-endpoint")

    def test_ui_contains_tap_focus_and_edit_notes_controls(self):
        response = self.client.get("/")
        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn('id="reader"', html)
        self.assertIn('id="editNotesButton"', html)
        self.assertIn('id="cameraZoomControls"', html)
        self.assertIn('id="zoomRange"', html)
        with open(os.path.join(os.path.dirname(application.__file__), "static", "app.js"), encoding="utf-8") as script:
            self.assertIn("Jika terbaca, aset akan dicari otomatis.", script.read())

    @patch("app.get_worksheet")
    def test_total_assets_survives_incomplete_log_headers(self, get_worksheet):
        self.log.values = [["TIMESTAMP", "NOMOR ASSET"]]
        get_worksheet.side_effect = self.worksheet
        _, headers = self.login()
        response = self.client.get("/api/dashboard", headers=headers)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json["summary"]["total"], 1)
        self.assertEqual(response.json["summary"]["pending"], 1)
        self.assertTrue(response.json["warnings"])

    @patch("app.get_worksheet")
    def test_dashboard_period_filter(self, get_worksheet):
        get_worksheet.side_effect = self.worksheet
        _, headers = self.login()
        self.log.values.append(["2026-01-15 10:00:00", "AST-0001", "LAPTOP", "LT-01", "BUDI", "DONE", "OK", "KANTOR", "AREA A", "BAIK", "SUDAH OPNAME", "2026-01-15 10:00:00", "", "ADMIN PIC", "1002", "SUPER ADMIN, PIC ASET"])
        january = self.client.get("/api/dashboard?startDate=2026-01-01&endDate=2026-01-31", headers=headers).json
        february = self.client.get("/api/dashboard?startDate=2026-02-01&endDate=2026-02-28", headers=headers).json
        self.assertEqual(january["summary"]["total"], 1)
        self.assertEqual(january["summary"]["completed"], 1)
        self.assertEqual(february["summary"]["total"], 1)
        self.assertEqual(february["summary"]["completed"], 0)

    @patch("app.get_worksheet")
    def test_scorecard_period_and_information_counts(self, get_worksheet):
        get_worksheet.side_effect = self.worksheet
        _, headers = self.login()
        self.log.values.append([
            "2026-01-15 10:00:00", "AST-0001", "LAPTOP", "LT-01", "BUDI", "DONE", "OK", "KANTOR",
            "AREA A", "BAIK", "SUDAH OPNAME", "2026-01-15 10:00:00", "Lengkap",
            "https://drive.google.com/file/d/1/view", "ADMIN PIC", "1002", "SUPER ADMIN, PIC ASET"
        ])
        jan_jun = self.client.get("/api/dashboard?scorePeriod=JAN-JUN", headers=headers).json
        jul_des = self.client.get("/api/dashboard?scorePeriod=JUL-DES", headers=headers).json
        jan_score = next(item for item in jan_jun["scoreCard"] if item["name"] == "ADMIN PIC")
        jul_score = next(item for item in jul_des["scoreCard"] if item["name"] == "ADMIN PIC")
        self.assertEqual(jan_score["progress"], 100)
        self.assertEqual(jan_score["documentationCount"], 1)
        self.assertEqual(jan_score["notesCount"], 1)
        self.assertEqual(jan_score["good"], 1)
        self.assertEqual(jul_score["progress"], 0)

    @patch("app.get_worksheet")
    def test_data_quality_summary_detail_period_and_export(self, get_worksheet):
        self.master.values.extend([
            values_for(application.MASTER_HEADERS, {"NOMOR ASSET": "AST-DUP", "TYPE": "MEJA", "USER": "A", "AREA": "AREA A", "LOKASI DETAIL": "LT 1"}),
            values_for(application.MASTER_HEADERS, {"NOMOR ASSET": "AST-DUP", "TYPE": "MEJA", "USER": "A", "AREA": "AREA A", "LOKASI DETAIL": "LT 2"}),
            values_for(application.MASTER_HEADERS, {"NOMOR ASSET": "BELUM ADA NOMOR ASSET", "TYPE": "", "USER": "", "AREA": "", "LOKASI DETAIL": ""}),
            values_for(application.MASTER_HEADERS, {"NOMOR ASSET": "AST-NOT", "TYPE": "KURSI", "USER": "B", "AREA": "AREA B", "LOKASI DETAIL": "LT 3"}),
        ])
        self.log.values.append(values_for(application.LOG_HEADERS, {
            "TIMESTAMP": "2026-01-15 10:00:00", "NOMOR ASSET": "AST-0001", "TYPE": "LAPTOP", "USER": "BUDI",
            "OPNAME": "DONE", "KONDISI": "RUSAK", "LOKASI DETAIL": "KANTOR", "AREA": "AREA A",
            "KONDISI TERAKHIR": "RUSAK", "STATUS TERAKHIR": "SUDAH OPNAME", "TANGGAL OPNAME TERAKHIR": "2026-01-15 10:00:00",
            "KETERANGAN TERAKHIR": "", "DOKUMENTASI TERAKHIR": "",
        }))
        self.log.values.append(values_for(application.LOG_HEADERS, {
            "TIMESTAMP": "2026-02-15 10:00:00", "NOMOR ASSET": "AST-DUP", "TYPE": "MEJA", "USER": "A",
            "OPNAME": "DONE", "KONDISI": "RUSAK", "LOKASI DETAIL": "LT 1", "AREA": "AREA A",
            "KONDISI TERAKHIR": "RUSAK", "STATUS TERAKHIR": "SUDAH OPNAME", "TANGGAL OPNAME TERAKHIR": "2026-02-15 10:00:00",
            "KETERANGAN TERAKHIR": "Perlu penggantian", "DOKUMENTASI TERAKHIR": "https://drive.google.com/file/d/1/view",
        }))
        get_worksheet.side_effect = self.worksheet
        _, headers = self.login("ADMIN", "1001")

        summary = self.client.get("/api/data-quality?period=JAN-JUN", headers=headers)
        self.assertEqual(summary.status_code, 200)
        counts = {item["key"]: item["count"] for item in summary.json["summary"]}
        self.assertEqual(counts["duplicate_asset"], 2)
        self.assertEqual(counts["missing_asset_number"], 1)
        self.assertEqual(counts["empty_area"], 1)
        self.assertEqual(counts["empty_location"], 1)
        self.assertEqual(counts["empty_type"], 1)
        self.assertEqual(counts["empty_user"], 1)
        self.assertEqual(counts["empty_documentation"], 1)
        self.assertEqual(counts["damaged_without_notes"], 1)
        self.assertEqual(counts["completed_opname"], 2)
        self.assertEqual(counts["damaged_with_notes"], 1)

        jul_des = self.client.get("/api/data-quality?period=JUL-DES", headers=headers).json
        jul_counts = {item["key"]: item["count"] for item in jul_des["summary"]}
        self.assertEqual(jul_counts["empty_documentation"], 0)
        self.assertEqual(jul_counts["damaged_without_notes"], 0)
        self.assertEqual(jul_counts["completed_opname"], 0)
        self.assertEqual(jul_counts["damaged_with_notes"], 0)

        detail = self.client.get("/api/data-quality/detail?category=duplicate_asset&period=ALL", headers=headers)
        self.assertEqual(detail.status_code, 200)
        self.assertEqual(len(detail.json["rows"]), 2)
        self.assertEqual(detail.json["rows"][0]["SUMBER"], "MASTER_ASET")
        self.assertTrue(detail.json["rows"][0]["BARIS"])

        exported = self.client.get("/api/data-quality/export?category=duplicate_asset&period=ALL&format=XLSX", headers=headers)
        self.assertEqual(exported.status_code, 200)
        self.assertEqual(exported.data[:2], b"PK")
        workbook = load_workbook(io.BytesIO(exported.data))
        self.assertTrue(workbook.active["A3"].alignment.wrap_text)

        pdf = self.client.get("/api/data-quality/export?category=damaged_with_notes&period=JAN-JUN&format=PDF", headers=headers)
        self.assertEqual(pdf.status_code, 200)
        self.assertEqual(pdf.data[:4], b"%PDF")

        ppt = self.client.get("/api/data-quality/export?category=damaged_with_notes&period=JAN-JUN&format=PPTX", headers=headers)
        self.assertEqual(ppt.status_code, 200)
        self.assertEqual(ppt.data[:2], b"PK")

        all_excel = self.client.get("/api/data-quality/export-all?period=JAN-JUN&format=XLSX", headers=headers)
        self.assertEqual(all_excel.status_code, 200)
        self.assertEqual(all_excel.data[:2], b"PK")
        all_workbook = load_workbook(io.BytesIO(all_excel.data))
        self.assertGreater(len(all_workbook.sheetnames), 1)

        all_pdf = self.client.get("/api/data-quality/export-all?period=JAN-JUN&format=PDF", headers=headers)
        self.assertEqual(all_pdf.status_code, 200)
        self.assertEqual(all_pdf.data[:4], b"%PDF")

        all_ppt = self.client.get("/api/data-quality/export-all?period=JAN-JUN&format=PPTX", headers=headers)
        self.assertEqual(all_ppt.status_code, 200)
        self.assertEqual(all_ppt.data[:2], b"PK")

    @patch("app.get_worksheet")
    def test_data_quality_access_and_incomplete_log_warning(self, get_worksheet):
        get_worksheet.side_effect = self.worksheet
        _, pic_headers = self.login("PIC MULTI", "1003")
        denied = self.client.get("/api/data-quality", headers=pic_headers)
        self.assertEqual(denied.status_code, 403)

        self.log.values = [["TIMESTAMP", "NOMOR ASSET"]]
        _, admin_headers = self.login("ADMIN", "1001")
        response = self.client.get("/api/data-quality", headers=admin_headers)
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json["warnings"])
        counts = {item["key"]: item["count"] for item in response.json["summary"]}
        self.assertEqual(counts["empty_documentation"], 0)
        self.assertEqual(counts["not_opnamed"], 0)

    @patch("app.get_worksheet")
    def test_asset_detail_still_loads_when_log_headers_incomplete(self, get_worksheet):
        self.log.values = [["TIMESTAMP", "NOMOR ASSET"]]
        get_worksheet.side_effect = self.worksheet
        _, headers = self.login()
        response = self.client.get("/api/assets/AST-0001", headers=headers)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json["asset"]["assetCode"], "AST-0001")
        self.assertEqual(response.json["history"], [])
        self.assertTrue(response.json["warnings"])

    @patch("app.get_worksheet")
    def test_submit_repairs_log_then_updates_master_opname_and_condition(self, get_worksheet):
        self.log.values = [["TIMESTAMP", "NOMOR ASSET"], ["OLD", "AST-OLD"]]
        get_worksheet.side_effect = self.worksheet
        _, headers = self.login()
        response = self.client.post("/api/opname", headers=headers, json={"assetCode": "AST-0001", "condition": "RUSAK", "notes": "Roda rusak"})
        self.assertEqual(response.status_code, 200)
        self.assertTrue(all(header in self.log.values[0] for header in application.LOG_HEADERS))
        appended = dict(zip(self.log.values[0], self.log.appended[0]))
        self.assertEqual(appended["OPNAME"], "DONE")
        self.assertEqual(appended["KONDISI"], "RUSAK")
        self.assertEqual(appended["KONDISI TERAKHIR"], "RUSAK")
        master_updates = {item["range"]: item["values"][0][0] for item in self.master.updates}
        self.assertEqual(master_updates["E2"], "DONE")
        self.assertEqual(master_updates["F2"], "RUSAK")
        self.assertIn(application.current_period_status(), master_updates.values())


if __name__ == "__main__":
    unittest.main()
